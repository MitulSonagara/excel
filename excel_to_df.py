import pandas as pd
from openpyxl import load_workbook


def abc(path):
    wb = load_workbook(path)
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    data = []

    for row in sheet.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True
    ):
        processed_row = []
        for cell_value in row:
            if cell_value is None:
                cell_value = processed_row[-1] if processed_row else None
            processed_row.append(cell_value)
        data.append(processed_row)

    df = pd.DataFrame(data)

    for merged_cell in sheet.merged_cells.ranges:
        if merged_cell.min_row != merged_cell.max_row:
            top_left_value = sheet.cell(
                row=merged_cell.min_row, column=merged_cell.min_col
            ).value
            for row_index in range(merged_cell.min_row + 1, merged_cell.max_row + 1):
                df.iloc[row_index - 1, merged_cell.min_col - 1] = top_left_value

    return df


df1 = abc("emergency.xlsx")
df1 = df1.drop(0)
df1 = df1.drop(df1.columns[[0, 1, 2, 6, 7, 8]], axis=1)
df1 = df1.reset_index(drop=True)
df1.columns = range(len(df1.columns))
df1[[0, 2]] = df1[[2, 0]]

grouped_df1 = (
    df1.groupby([0, 1])
    .agg({2: lambda x: "\n".join(map(str, x))})
    .reset_index()
)

# Adding the 'total' column to grouped_df1
grouped_df1[3] = df1.groupby([0, 1]).size().values

# Correcting the column assignments
grouped_df1[[2,3]] = grouped_df1[[3, 2]]


co = [4,5,6,7,8]
for i in range(5):
    grouped_df1.insert(loc=3 + i, column=co[i], value=None)
grouped_df1.columns = range(len(grouped_df1.columns))

new_row = pd.DataFrame(
    {
        0: ["BLOCK NO"],
        1: ["SUBJECT CODE"],
        2: ["TOTAL PRESENT STUDENTS (A)"],
        3:["NO OF ABSENT STUDENTS (B)"],
        4:["UFM CASE(C)"],
        5:["TOTAL (A+B+C)"],
        6:["SEAT NO OF ABSENT STUDENTS"],
        7:["SEAT NO OF UFM CASES"],
        8: ["EMERGENCY STUDENTS IF ANY"],
        
    }
)
grouped_df1 = pd.concat([new_row, grouped_df1]).reset_index(drop=True)

print(grouped_df1)
