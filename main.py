import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

def excel_to_df(path):
    wb = load_workbook(path)

    # Select the first ws (you can choose a specific sheet if needed)
    sheet = wb.active

    # Get the max row and max column count from the sheet
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Initialize an empty list to hold the data
    data = []

    for row in sheet.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True
    ):
        # Create a list to hold processed row data
        processed_row = []
        for cell_value in row:
            if cell_value is None:
                # If cell is None, propagate value from the top-left cell of the merged region
                cell_value = processed_row[-1] if processed_row else None
            processed_row.append(cell_value)
        data.append(processed_row)

    df = pd.DataFrame(data)

    for merged_cell in sheet.merged_cells.ranges:
        # Check if the merged cell range is vertical
        if merged_cell.min_row != merged_cell.max_row:
            # Get the top-left cell value
            top_left_value = sheet.cell(
                row=merged_cell.min_row, column=merged_cell.min_col
            ).value
            # Assign the top-left cell value to all cells within the merged range vertically
            for row_index in range(merged_cell.min_row + 1, merged_cell.max_row + 1):
                df.iloc[row_index - 1, merged_cell.min_col - 1] = top_left_value

    return df

def process_df1(df):
    df = df.drop([0, 1, 2, 3, 4])
    df = df.drop(df.columns[[2, 3, 4]], axis=1)

    # Reset the index of the DataFrame after dropping rows and columns
    df = df.reset_index(drop=True)
    df.columns = range(len(df.columns))

    # Swap columns 0 and 1
    df[[0, 1]] = df[[1, 0]]

    co = [3, 4, 5, 6, 7, 8]
    for i in range(6):
        df.insert(loc=3 + i, column=co[i], value=None)
    df.columns = range(len(df.columns))
    df = df.drop([0, 34])
    return df

def process_df2(df):
    df = df.drop(0)
    df = df.drop(df.columns[[0, 1, 2, 6, 7, 8]], axis=1)
    df = df.reset_index(drop=True)
    df.columns = range(len(df.columns))
    df[[0, 2]] = df[[2, 0]]

    df2 = df.groupby([0, 1]).agg({2: lambda x: "\n".join(map(str, x))}).reset_index()

    # Adding the 'total' column to df2
    df2[3] = df.groupby([0, 1]).size().values

    # Correcting the column assignments
    df2[[2, 3]] = df2[[3, 2]]

    co = [4, 5, 6, 7, 8]
    for i in range(5):
        df2.insert(loc=3 + i, column=co[i], value=None)
    df2.columns = range(len(df2.columns))

    def modify_block_no(row):
        return str(row) + '/E'

    df2[0] = df2[0].apply(modify_block_no)

    return df2

df1 = excel_to_df("input1.xlsx")
df1 = process_df1(df1)
df1 = pd.DataFrame(df1)

df2 = excel_to_df("emergency.xlsx")
df2 = process_df2(df2)
df2 = pd.DataFrame(df2)

df1[1] = df1[1].astype("int64")
df2[1] = df2[1].astype("int64")

current_subject_code = None
new_rows = []
for index, row in df1.iterrows():
    if row[1] != current_subject_code:
        if current_subject_code is not None:
            for index, row1 in df2.iterrows():
                subject = row1[1]
                if current_subject_code == subject:
                    new_rows.append(row1)
                    df2 = df2.drop(index)
                    df2 = df2.reset_index(drop=True)
        current_subject_code = row[1]
    new_rows.append(row)


df = pd.DataFrame(new_rows)
df = df.reset_index(drop=True)
df = pd.concat([df, df2], ignore_index=True)
df[[3,4]]=0

# Insert a blank row after each group of subject code
current_subject_code = None
new_rows = []
for index, row in df.iterrows():
    if row[1] != current_subject_code:
        if current_subject_code is not None:
            new_rows.append(pd.Series([None] * len(df.columns), index=df.columns))
            new_rows[-1][0] = "Total"
            new_rows[-1][1] = None
        current_subject_code = row[1]
    new_rows.append(row)

new_rows.append(pd.Series([None] * len(df.columns), index=df.columns))
new_rows[-1][0] = "Total"
new_rows[-1][1] = None

# Create a new DataFrame with inserted blank rows
df_with_blank_rows = pd.DataFrame(new_rows)
# Reset the index of the DataFrame
df_with_blank_rows = df_with_blank_rows.reset_index(drop=True)

cols = [
    "BLOCK\nNO.",
    "SUBJECT\nCODE",
    "NO OF\nPRESENT\nSTUDENTS\n(A)",
    "NO OF\nABSENT\nSTUDENTS\n(B)",
    "UFM\nCASE\n(C)",
    "TOTAL\n(A + B + C)",
    "SEAT NO\nOF ABSENT\nSTUDENTS",
    "SEAT NO\nOF UFM\nCASES",
    "EMERGENCY\nSTUDENTS\nIF ANY",
]
# Manually set the column names
df_with_blank_rows.columns = cols

output_excel_file = "output1.xlsx"
df_with_blank_rows.to_excel(output_excel_file, index=False)
excel_file = "output1.xlsx"  # Replace "input1.xlsx" with the path to your Excel file
wb1 = load_workbook(excel_file)
# Select the first sheet (you can choose a specific sheet if needed)
ws = wb1.active

# Define the desired widths for each column
column_widths = {"A": 8, "B": 15, "C": 12,"D":12,"E":6,"F":12,"G":32,"H":10,"I":21}  # Example widths for columns A, B, and C

# Set the width of columns based on the defined widths
for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Add borders to all cells
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


ws.move_range("A1:{}{}".format(get_column_letter(ws.max_column), ws.max_row), rows=6)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ws.max_column)
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ws.max_column)
ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=ws.max_column)
ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ws.max_column)


# Write data to the new row
ws["A1"] = "GUJARAT TECHNOLOGICAL UNIVERSITY"
ws["A2"] = "AHMEDABAD - GANDHINAGAR"
ws["A3"] = "DATE : "
ws["G3"] = "SESSION : "
ws["A4"] = "COLLEGE CODE: 017 "
ws["A5"] = "COLLEGE NAME: Vishwakarama Government Engineering College, Chandkheda "
# Add more columns as needed

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    # Check if the first cell is "Total" and the second cell is empty
    if row[0].value == "Total" and not row[1].value:
        # Merge cells for "Total"
        ws.merge_cells(
            start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=2
        )

for row in ws.iter_rows(
    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
):
    for cell in row:
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )

total = []
start_row = 2
end_row = None
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
for row_num in range(8, ws.max_row + 1):
    # Check if the cell in column A is "Total"
    if ws[f"A{row_num}"].value == "Total":
        total.append(row_num)
        # Apply yellow fill to the entire row
        for col_num in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_num).fill = yellow_fill

        # Calculate sums and update cells
        ws[f"F{row_num}"] = f"=SUM(F{start_row}:F{end_row})"
        ws[f"E{row_num}"] = f"=SUM(E{start_row}:E{end_row})"
        ws[f"D{row_num}"] = f"=SUM(D{start_row}:D{end_row})"
        ws[f"C{row_num}"] = f"=SUM(C{start_row}:C{end_row})"

        # Update start_row for the next section
        start_row = row_num + 1
    else:
        # Update end_row if the value in column F is not 0
        if ws[f"F{row_num}"].value != 0:
            end_row = row_num

print(total)

for row_num in range(8, ws.max_row + 1):
    ws[f"F{row_num}"] = f"=SUM(C{row_num}:E{row_num})"

last_row = ws.max_row
new_row = last_row + 1

ws.merge_cells(start_row=new_row, start_column=1, end_row=new_row, end_column=2)
ws[f"A{new_row}"] = "GRAND TOTAL"

for char in ("C","D","E","F"):
    # Convert row numbers to Excel-style cell references (e.g., 15 -> 'C15')
    cell_references = [f"{char}{row_num}" for row_num in total]
    # Join the cell references with '+' to create the formula string
    formula_string = ",".join(cell_references)
    # Insert the formula in cell C66
    ws[f"{char}{new_row}"] = f"=SUM({formula_string})"


font_style = Font(name="Calibri Light", size=12,bold=True)

# Set the height of the first five rows to 16
for row_num in range(1, 7):
    ws.row_dimensions[row_num].height = 16

# Iterate through the first five rows and set the font style and size for each cell
for row in ws.iter_rows(min_row=1, max_row=5):
    for cell in row:
        cell.font = font_style

for row in ws.iter_rows(
    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
):
    for cell in row:
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )

for row_num in range(3, 7):  # Rows 3, 4, and 5
    for cell in ws[row_num]:  # Iterate through cells in the row
        cell.alignment = Alignment(horizontal="left")

ws["G3"].alignment = Alignment(horizontal="right")

for row in ws.iter_rows():
    for cell in row:
        cell.border = border

last_row = ws.max_row
for cell in ws[last_row]:
    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    cell.font = font_style

wb1.save(excel_file)

# wb1.save(excel_file)

# print(f"Excel file '{output_excel_file}' created successfully.")
